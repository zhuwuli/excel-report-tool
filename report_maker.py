#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel汇总表自动处理工具 + Excel转PDF 一条龙
流程：
  1. 复制目录（期数+1，日期+自定义天数）
  2. 删除目录下所有PDF文件
  3. 修改Excel封面（期数、天气、温度）
  4. 处理Excel日期列（列迁移只写数值）
  5. 自动处理外部引用（注册表+自动点击更新）
  6. Excel转PDF（自动处理公式重算）
"""

import os
import re
import shutil
import sys
import time
import subprocess
import json
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from functools import wraps
import warnings
import requests
import random
import win32com.client
import winreg
import pywinauto

# 共享工具：颜色输出、IDE环境检测
from utils import green, red, yellow, cyan, bold, HAS_COLOR

warnings.filterwarnings('ignore')

# ============ 可选依赖 ============
try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

# ============ Windows Console 进度条（绕过 stdout 管道缓冲） ============
import ctypes

STD_OUTPUT_HANDLE = -11

class WindowsConsoleProgress:
    """用 Windows Console API 画进度条，绕过 Python stdout 管道缓冲"""

    def __init__(self, total, desc="处理中"):
        self.total = total
        self.current = 0
        self.desc = desc
        self._handle = ctypes.windll.kernel32.GetStdHandle(STD_OUTPUT_HANDLE)
        self._done = False

    def _write(self, text):
        """直接写 Windows Console，绕过 Python stdout 管道缓冲"""
        ctypes.windll.kernel32.WriteConsoleW(self._handle, text, len(text), None, None)

    def _clear_line(self):
        """清除当前行（ANSI 清除到行尾）"""
        self._write('\r' + ' ' * 70 + '\r')

    def update(self, current, label=""):
        self.current = current
        pct = int(current / self.total * 20) if self.total > 0 else 0
        bar = '█' * pct + '░' * (20 - pct)
        label_text = f" {label}" if label else ""
        msg = f"  {self.desc}: {bar} {current}/{self.total}{label_text}"
        self._clear_line()
        self._write(msg)

    def finish(self):
        if not self._done:
            self._write('\r' + ' ' * 70 + '\r')
            self._done = True

HAS_WIN_CONSOLE = sys.stdout.isatty()  # 仅在真终端启用，管道模式退化

# ═══════════════════════════════════════════════════════════════════════════════════
# 使用者配置区（按需修改这里）
# ═══════════════════════════════════════════════════════════════════════════════════

# ── 基础路径 ──
BASE_DIR     = r"C:\Users\Lenovo\Desktop\工作文件"     # 工程根目录
FALLBACK_DIR = "第87期 04.09"                          # 默认起始文件夹

# ── Excel 类型 ──
#   "office" - Microsoft Office（默认，静默运行）
#   "wps"    - WPS Excel（有两级对话框，需 Visible=True）
EXCEL_TYPE   = "office"
EXCEL_VISIBLE = (EXCEL_TYPE.lower() == "wps")           # 自动计算，无需修改

# ── 天气 API ──
WEATHER_API_KEY = "S-db8PJBJo-bZ0rb0"                  # 心知天气 Key，过期请更新
JINAN_LAT       = 36.6512                               # 济南坐标（纬度）
JINAN_LON       = 117.1205                              # 济南坐标（经度）

# ── 处理范围 ──
TARGET_SHEETS = {"垂直位移", "坑外水位", "水位", "测斜", "水平位移", "水平","竖向位移", "竖向","轴力"}  # 要处理的 sheet 名称

# ═══════════════════════════════════════════════════════════════════════════════════
# 内部配置（一般不需修改）
# ═══════════════════════════════════════════════════════════════════════════════════

# 天气关键词（用于封面匹配）
WEATHER_KEYWORDS = ["晴", "多云", "阴", "雨","小雨", "中雨", "大雨", "雷阵雨", "小雪", "大雪", "雾", "霾"]

# 季节天气/温度配置
SEASON_WEATHER = {
    (3, 4, 5): ["晴", "多云", "阴", "小雨"],
    (6, 7, 8): ["晴", "多云", "雷阵雨", "大雨"],
    (9, 10, 11): ["晴", "多云", "阴", "小雨"],
    (12, 1, 2): ["晴", "多云", "阴", "小雪"],
}
SEASON_TEMP = {
    (12, 1, 2): (-5, 5),
    (3, 4, 5): (10, 20),
    (6, 7, 8): (25, 35),
    (9, 10, 11): (15, 25),
}

# 天气代码映射
WEATHER_CODE_MAP = {                                    # 心知天气代码
    "0": "晴", "1": "多云", "2": "阴", "3": "小雨", "4": "中雨",
    "5": "大雨", "6": "雷阵雨", "7": "小雪", "8": "中雪", "9": "大雪",
    "10": "雾", "13": "阵雨", "14": "雷阵雨", "19": "霾", "23": "雾",
}
WMO_CODE_MAP = {                                        # Open-Meteo WMO 代码
    0: "晴", 1: "晴", 2: "多云", 3: "阴",
    45: "雾", 48: "雾", 51: "小雨", 53: "小雨", 55: "小雨",
    61: "雨", 63: "雨", 65: "雨",
    80: "阵雨", 81: "阵雨", 82: "阵雨",
    95: "雷阵雨", 96: "雷阵雨", 99: "雷阵雨"
}

# 重试配置
MAX_RETRIES    = 3       # 各操作最大重试次数
RETRY_BACKOFF   = [1, 2, 4]   # 指数退避间隔（秒）

# Step5 智能等待配置
STEP5_STARTUP_MIN_WAIT = 0.5
STEP5_STARTUP_TIMEOUT = 2.0
STEP5_CALC_MIN_WAIT = 0.5
STEP5_CALC_TIMEOUT = 5.0
STEP5_DIALOG_TIMEOUT = {"office": 6.0, "wps": 8.0}

# Excel/WPS 临时锁文件
TEMP_LOCK_PREFIX = '~$'
TEMP_EXCEL_EXTS = ('.xls', '.xlsx', '.xlsm')


class PipelineCancelled(Exception):
    """用户请求安全停止流水线"""


def check_cancelled(cancel_check=None):
    """在步骤边界检查是否请求取消"""
    if cancel_check and cancel_check():
        raise PipelineCancelled("用户请求停止处理")

# ============ 通用装饰器 ============
def with_retry(func):
    """
    通用重试装饰器：失败自动重试 MAX_RETRIES 次，指数退避
    适用于天气 API 等可能临时失败的网络请求
    """
    @wraps(func)
    def wrapper(*args, **kwargs):
        for i in range(MAX_RETRIES):
            try:
                return func(*args, **kwargs)
            except Exception:
                if i < MAX_RETRIES - 1:
                    time.sleep(RETRY_BACKOFF[i])
                else:
                    return None
    return wrapper


# ============ 外部链接处理 ============
def add_trusted_location(folder_path):
    """将目录添加到Excel信任位置（注册表），避免黄色警告条"""
    try:
        key_path = r"Software\Microsoft\Office\16.0\Excel\Security\Trusted Locations"
        key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, key_path)
        sub_key_name = f"Location{abs(hash(folder_path)) % 100000}"
        sub_key = winreg.CreateKey(key, sub_key_name)
        winreg.SetValueEx(sub_key, "Path", 0, winreg.REG_SZ, folder_path)
        winreg.SetValueEx(sub_key, "AllowSubFolders", 0, winreg.REG_DWORD, 1)
        winreg.CloseKey(sub_key)
        winreg.CloseKey(key)
        print(f"  [成功] 已添加信任位置: {folder_path}")
        return True
    except Exception as e:
        print(f"  [失败] 添加信任位置失败: {e}")
        return False

def _excel_process_names():
    """根据当前 Excel 类型返回需要清理的进程名"""
    names = ["EXCEL.EXE"]
    if EXCEL_TYPE.lower() == "wps":
        names.append("et.exe")
    return names


def _kill_excel_process(process_names=None):
    """强制关闭 Excel/WPS 表格进程，确保 COM 实例完全释放"""
    names = process_names or _excel_process_names()
    for name in names:
        try:
            subprocess.run(["taskkill", "/F", "/IM", name],
                           capture_output=True, timeout=5)
        except Exception:
            pass
    time.sleep(1.5)  # 等待进程完全退出

_desktop = None
try:
    _desktop = pywinauto.Desktop(backend="win32")
except Exception:
    pass

def _wait_for_excel_ready(excel, cancel_check=None):
    """至少等待短暂启动缓冲，并在上限内等待 Excel 就绪。"""
    start = time.monotonic()
    deadline = start + STEP5_STARTUP_TIMEOUT
    while True:
        check_cancelled(cancel_check)
        elapsed = time.monotonic() - start
        try:
            ready = bool(excel.Ready)
        except Exception:
            ready = True
        if elapsed >= STEP5_STARTUP_MIN_WAIT and ready:
            return True
        if time.monotonic() >= deadline:
            return False
        time.sleep(0.1)


def _wait_for_excel_calculation(excel, cancel_check=None):
    """等待链接更新后的计算完成；无法读取状态时保留最小安全等待。"""
    start = time.monotonic()
    deadline = start + STEP5_CALC_TIMEOUT
    while True:
        check_cancelled(cancel_check)
        elapsed = time.monotonic() - start
        try:
            calculation_done = int(excel.CalculationState) == 0
            ready = bool(excel.Ready)
        except Exception:
            calculation_done = True
            ready = True
        if elapsed >= STEP5_CALC_MIN_WAIT and calculation_done and ready:
            return True
        if time.monotonic() >= deadline:
            return False
        time.sleep(0.1)


def _workbook_has_external_links(workbook):
    """返回是否有 Excel 外部链接；接口不兼容时返回 None。"""
    try:
        return bool(workbook.LinkSources(1))
    except Exception:
        return None


def _click_excel_update_button(timeout=None, cancel_check=None):
    """用pywinauto点击Excel'是否更新链接'对话框的'更新'按钮。"""
    if _desktop is None:
        return False

    if timeout is None:
        timeout = STEP5_DIALOG_TIMEOUT.get(EXCEL_TYPE.lower(), 6.0)

    def try_click():
        try:
            for win in _desktop.windows():
                if win.window_text() == "Microsoft Excel":
                    for child in win.children():
                        txt = child.window_text()
                        # 精确匹配"更新(&U)"按钮
                        if txt == "更新(&U)":
                            child.click()
                            return True
        except Exception:
            pass
        return False

    deadline = time.monotonic() + timeout
    while True:
        check_cancelled(cancel_check)
        if try_click():
            return True
        if time.monotonic() >= deadline:
            return False
        time.sleep(0.25)

def open_and_update_links(excel_path, max_retries=MAX_RETRIES, cancel_check=None):
    """
    用COM打开Excel文件，自动处理'是否更新外部链接'对话框。
    流程：打开文件（UpdateLinks=2）→ 等对话框弹出 → pywinauto点'更新' → 保存关闭
    失败时最多重试 max_retries 次（指数退避）
    """
    last_error = None

    for attempt in range(max_retries):
        check_cancelled(cancel_check)
        def _do():
            started_at = time.monotonic()
            check_cancelled(cancel_check)
            _kill_excel_process()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = EXCEL_VISIBLE
            print(f"[DEBUG Step5] excel_type={EXCEL_TYPE}, EXCEL_VISIBLE={EXCEL_VISIBLE}, Visible set to {EXCEL_VISIBLE}")
            excel.DisplayAlerts = False
            if not _wait_for_excel_ready(excel, cancel_check):
                print("  [提示] Excel 启动等待达到上限，继续尝试打开文件")

            wb = excel.Workbooks.Open(excel_path, UpdateLinks=2)
            has_external_links = _workbook_has_external_links(wb)
            if has_external_links is False:
                print("  [提示] 工作簿无外部链接，跳过更新按钮等待")
            else:
                clicked = _click_excel_update_button(cancel_check=cancel_check)
                if clicked:
                    print(f"  [成功] 自动点击了'更新'按钮")
                else:
                    print(f"  [提示] 未检测到链接更新对话框")

            if not _wait_for_excel_calculation(excel, cancel_check):
                print("  [提示] Excel 计算等待达到上限，继续保存")

            wb.Save()
            wb.Close()
            excel.Quit()
            _kill_excel_process()
            elapsed = time.monotonic() - started_at
            print(f"  [耗时] Step5 单文件 {elapsed:.1f}s")
            return True

        try:
            if attempt > 0:
                wait_seconds = RETRY_BACKOFF[attempt - 1]
                print(f"  [重试] 第 {attempt + 1} 次尝试，等待 {wait_seconds}s...")
                time.sleep(wait_seconds)
            result = _do()
            if result:
                return True
        except PipelineCancelled:
            _kill_excel_process()
            raise
        except Exception as e:
            last_error = e
            print(f"  [失败] 第 {attempt + 1} 次: {e}")
            _kill_excel_process()

    print(f"  [最终] {os.path.basename(excel_path)} {max_retries}次重试均失败: {last_error}")
    return False

def handle_external_links_auto(target_dir, cancel_check=None):
    """
    自动处理目录下的外部引用：
    1. 将目录加入Excel信任位置（避免黄色警告条）
    2. 遍历所有Excel文件，每个文件单独开一个进程处理
    """
    print("\n[工具] 步骤5: 自动处理外部引用")

    # 1. 添加信任位置
    print("  [注册] 添加Excel信任位置...")
    add_trusted_location(target_dir)

# 2. 遍历所有Excel文件，自动处理外部链接
    all_excels = find_all_excel_files(target_dir)
    success_count = 0
    fail_count = 0

    # 用 Windows Console API 进度条，绕过 stdout 管道缓冲（实时更新）
    if HAS_WIN_CONSOLE:
        progress = WindowsConsoleProgress(len(all_excels), desc="Step5处理")
    else:
        progress = None

    for i, excel_path in enumerate(all_excels, 1):
        check_cancelled(cancel_check)
        basename = os.path.basename(excel_path)
        if progress:
            progress.update(i, basename)
        else:
            print(f"  [处理] {basename}...", flush=True)

        if open_and_update_links(excel_path, cancel_check=cancel_check):
            success_count += 1
            print(f"  {green('[成功]')} {basename} 处理完成")
        else:
            fail_count += 1
            print(f"  {red('[失败]')} {basename} 处理失败（已重试{MAX_RETRIES}次）")
        time.sleep(0.2)

    if progress:
        progress.finish()
        print()  # 换行，避免后续输出覆盖同一行

    print(f"\n{cyan('[完成]')} 外部引用处理完成: {success_count}/{len(all_excels)} 个文件")
    if fail_count > 0:
        print(f"  {yellow('[警告]')} {fail_count} 个文件处理失败")
    return success_count



# ============ 工具函数 ============
def parse_folder_info(name):
    """解析文件夹名称，提取期数和日期"""
    m = re.search(r'第(\d+)期\s+(\d{2})\.(\d{2})', name)
    return (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else (None, None, None)

def calculate_new_date(folder_name, year, days):
    """计算新日期（期数+1，日期+days天）"""
    issue, month, day = parse_folder_info(folder_name)
    if issue is None:
        return None, None
    new_date = datetime(year, month, day) + timedelta(days=days)
    return f"第{issue + 1}期 {new_date.month:02d}.{new_date.day:02d}", new_date

def get_year_from_excel(path):
    """从Excel提取年份"""
    wb = None
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for col in range(1, 11):
            cell = ws.cell(row=1, column=col)
            if is_date_cell(cell):
                if isinstance(cell.value, datetime):
                    return cell.value.year
                else:
                    return datetime.strptime(str(cell.value), '%Y/%m/%d').year
        print("[警告] 未找到日期，使用当前年份")
        return None
    except Exception as e:
        print(f"[失败] 读取Excel失败: {e}")
        return None
    finally:
        if wb is not None:
            wb.close()

def get_weather(target_date):
    """获取天气（API优先，失败则模拟）"""
    days_diff = (target_date.date() - datetime.now().date()).days
    print(f"[搜索] 目标日期: {target_date.strftime('%Y-%m-%d')} (距今天 {days_diff} 天)")

    if days_diff < 0:
        # 过去日期，用 Historical API 查历史天气
        print(f"[调试] 过去日期，调用 Open-Meteo Historical...")
        w, t = fetch_open_meteo_historical(target_date)
        if w:
            print(f"[调试] Historical 返回: {w} {t}")
            return f"{w} {t}"
        print(f"[调试] Historical 失败，尝试 Open-Meteo Forecast...")
        w, t = fetch_open_meteo_weather()
        if w:
            return f"{w} {t}"
    elif days_diff <= 16:
        print(f"[调试] 尝试调用心知天气 API...")
        weather = fetch_api_weather(target_date, days_diff)
        if weather:
            print(f"[调试] 获取到天气: {weather}")
            return weather
    else:
        print(f"[调试] 日期超出 16 天范围，使用模拟天气")

    print("[切换] 使用模拟天气...")
    return make_simulated_weather(target_date)

@with_retry
def fetch_open_meteo_weather():
    """Open-Meteo Forecast API（备用，济南固定坐标）"""
    url = (f"https://api.open-meteo.com/v1/forecast"
           f"?latitude={JINAN_LAT}&longitude={JINAN_LON}"
           f"&daily=weathercode,temperature_2m_max,temperature_2m_min"
           f"&timezone=Asia/Shanghai&forecast_days=1")

    with urllib.request.urlopen(url, timeout=5) as r:
        data = json.loads(r.read().decode())

    code = data["daily"]["weathercode"][0]
    temp_max = data["daily"]["temperature_2m_max"][0]
    temp_min = data["daily"]["temperature_2m_min"][0]

    weather = WMO_CODE_MAP.get(code, "多云")
    temperature = f"{int(temp_min)}~{int(temp_max)}°C"
    return weather, temperature

@with_retry
def fetch_open_meteo_historical(target_date):
    """Open-Meteo Historical API 获取历史天气（济南）"""
    date_str = target_date.strftime('%Y-%m-%d')
    url = (f"https://archive-api.open-meteo.com/v1/archive"
           f"?latitude={JINAN_LAT}&longitude={JINAN_LON}"
           f"&start_date={date_str}&end_date={date_str}"
           f"&daily=weathercode,temperature_2m_max,temperature_2m_min"
           f"&timezone=Asia/Shanghai")

    with urllib.request.urlopen(url, timeout=10) as r:
        data = json.loads(r.read().decode())

    daily = data.get("daily", {})
    code = daily.get("weathercode", [None])[0]
    temp_max = daily.get("temperature_2m_max", [None])[0]
    temp_min = daily.get("temperature_2m_min", [None])[0]

    if code is None or temp_max is None:
        return None, None

    weather = WMO_CODE_MAP.get(code, "多云")
    temperature = f"{int(temp_min)}~{int(temp_max)}°C"
    print(f"  [历史] Open-Meteo Historical: {weather} {temperature}")
    return weather, temperature

@with_retry
def _fetch_seniverse_weather(target_date, days_diff):
    """从心知天气API获取天气（济南）"""
    url = (f"https://api.seniverse.com/v3/weather/daily.json"
           f"?key={WEATHER_API_KEY}&location=jinan&language=zh-Hans&unit=c"
           f"&days={days_diff + 1}")
    resp = requests.get(url, timeout=10)
    if resp.status_code != 200:
        raise Exception(f"HTTP {resp.status_code}")  # 触发重试

    data = resp.json()
    daily = data.get("results", [{}])[0].get("daily", [])
    target_str = target_date.strftime('%Y-%m-%d')

    for day in daily:
        if day.get("date") == target_str:
            text = day.get("text_day") or WEATHER_CODE_MAP.get(day.get("code_day", "1"), "多云")
            low, high = day.get("low", "15"), day.get("high", "25")
            result = f"{text} {low}~{high}°C"
            print(f"[成功] API获取成功: {result}")
            return result

    if daily:
        day = daily[0]
        text = WEATHER_CODE_MAP.get(day.get("code_day", "1"), "多云")
        low, high = day.get("low", "15"), day.get("high", "25")
        print(f"[警告] 日期匹配失败，使用最近可用数据")
        return f"{text} {low}~{high}°C"

def fetch_api_weather(target_date, days_diff):
    """获取天气，先试心知天气，失败则用 Open-Meteo 备用"""
    # 心知天气只能覆盖约7天内，超过7天直接用 Open-Meteo
    if days_diff <= 2:
        weather = _fetch_seniverse_weather(target_date, days_diff)
        if weather is not None:
            return weather

    # 超出心知天气范围或心知天气失败，尝试 Open-Meteo
    print(f"  [备用] 尝试 Open-Meteo...")
    w, t = fetch_open_meteo_weather()
    if w:
        print(f"  [备用] Open-Meteo 天气: {w}, 温度: {t}")
        return f"{w} {t}"
    return None

def make_simulated_weather(target_date):
    """生成模拟天气"""
    month = target_date.month
    for months, weather_list in SEASON_WEATHER.items():
        if month in months:
            weather = random.choice(weather_list)
            low, high = SEASON_TEMP[months]
            return f"{weather} {low}~{high}°C"
    return "多云 15~25°C"

def is_temp_excel_file(filename):
    """判断是否为 WPS/Excel 生成的临时锁文件"""
    name = os.path.basename(str(filename))
    return name.startswith(TEMP_LOCK_PREFIX) and name.lower().endswith(TEMP_EXCEL_EXTS)


def find_temp_excel_files(directory):
    """查找目录下残留的 WPS/Excel 临时锁文件"""
    temp_files = []
    for root, _, files_list in os.walk(directory):
        for f in files_list:
            if is_temp_excel_file(f):
                temp_files.append(os.path.join(root, f))
    return temp_files


def _ignore_temp_excel_files(_directory, names):
    """copytree 回调：复制目录时忽略临时锁文件"""
    return [name for name in names if is_temp_excel_file(name)]


def find_excel_files(directory, patterns=None):
    """查找Excel文件，排除WPS/Excel生成的临时锁文件"""
    patterns = patterns or ('汇总', 'summary')
    files = []
    for root, _, files_list in os.walk(directory):
        for f in files_list:
            if is_temp_excel_file(f):
                continue
            if f.lower().endswith(TEMP_EXCEL_EXTS):
                if any(p.lower() in f.lower() for p in patterns):
                    files.append(os.path.join(root, f))
    return files

def find_all_excel_files(directory):
    """查找目录下所有Excel文件（不过滤文件名，排除临时文件）"""
    return [
        os.path.join(directory, f)
        for f in os.listdir(directory)
        if f.lower().endswith(TEMP_EXCEL_EXTS)
        and not is_temp_excel_file(f)
    ]

def rename_excel_files_by_issue(directory, old_issue, new_issue):
    """根据期数重命名目录下的Excel文件（如 第87期.xlsx -> 第88期.xlsx）"""
    old_str = f"第{old_issue}期"
    new_str = f"第{new_issue}期"

    for root, _, files in os.walk(directory):
        for f in files:
            if old_str not in f:
                continue
            old_path = os.path.join(root, f)
            new_name = f.replace(old_str, new_str)
            new_path = os.path.join(root, new_name)
            try:
                os.rename(old_path, new_path)
                print(f"  [成功] 重命名: {f} -> {new_name}")
            except Exception as e:
                print(f"  [失败] 重命名失败: {f} -> {e}")

def is_date_cell(cell):
    """判断是否为日期单元格"""
    if not cell or cell.value is None:
        return False
    if isinstance(cell.value, datetime):
        return True
    fmt = str(cell.number_format).lower()
    return any(f in fmt for f in ['yy', 'yyyy', 'mm', 'dd', 'h', 'hh', 'mm:ss']) and \
        all(f not in fmt for f in ['$', '%', '0', '#', 'general'])

# ============ Excel处理 ============
def _range_to_column_values(value):
    """把 Excel Range.Value 统一转换为一维列数据列表"""
    if isinstance(value, tuple):
        values = []
        for item in value:
            if isinstance(item, tuple):
                values.append(item[0] if item else None)
            else:
                values.append(item)
        return values
    return [value]


def _column_values_to_range(values):
    """把一维列数据转换为 Excel Range.Value 需要的二维列形状"""
    return tuple((value,) for value in values)


def modify_cover(wb, issue_num, weather, date):
    """修改封面（期数、天气）"""
    try:
        ws = wb[wb.sheetnames[0]]
        weather_text = weather.split()[0]
        weather_ok = False

        def process_cell(cell):
            nonlocal weather_ok
            if cell.value is None:
                return
            val = str(cell.value)

            # 期数
            if "第" in val and "期" in val:
                m = re.search(r'第(\d+)期', val)
                if m:
                    print(f"  [调试] 找到期数单元格: {cell.coordinate} = {repr(cell.value)}")
                    new_val = re.sub(r'第\d+期', f'第{issue_num}期', val)
                    if new_val != val:
                        cell.value = new_val
                        print(f"  [成功] 修改期数: {cell.value}")

            # 天气关键词
            if not weather_ok:
                for kw in WEATHER_KEYWORDS:
                    if kw in val:
                        print(f"  [调试] 找到天气单元格: {cell.coordinate} = {repr(cell.value)}")
                        if "天气：" in val:
                            # 有前缀，保持前缀替换天气部分
                            cell.value = re.sub(r'天气：\S+', f'天气：{weather_text}', val)
                        else:
                            # 没有前缀，只替换天气关键词本身
                            cell.value = weather_text
                        print(f"  [成功] 修改天气: {cell.value}")
                        weather_ok = True
                        break

        # 封面区域搜索（1-20行，5-40列）
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=5, max_col=40):
            for cell in row:
                process_cell(cell)
                if weather_ok:
                    break

        # 全表搜索（期数和天气）
        if not weather_ok:
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=100):
                for cell in row:
                    process_cell(cell)

        print(f"[成功] 封面修改完成: 第{issue_num}期, 天气: {weather}")
        return True
    except Exception as e:
        print(f"[失败] 修改封面失败: {e}")
        return False

def process_excel(filepath, target_date, issue_num, weather, days=1, cancel_check=None):
    """处理汇总表 - 全部用COM完成"""
    if not os.path.exists(filepath):
        print(f"[失败] 文件不存在: {filepath}")
        return False

    excel_com = None
    wb_com = None
    try:
        check_cancelled(cancel_check)
        print(f"\n[文件] {os.path.basename(filepath)}")
        print(f"[日期] 目标日期: {target_date.strftime('%Y/%m/%d')}")

        _kill_excel_process()
        time.sleep(1.5)
        check_cancelled(cancel_check)
        excel_com = win32com.client.Dispatch("Excel.Application")
        excel_com.Visible = False
        excel_com.DisplayAlerts = False
        excel_com.ScreenUpdating = False
        wb_com = excel_com.Workbooks.Open(filepath)

        # ===== 步骤1: 修改封面信息 =====
        print("\n[工具] 步骤1: 修改封面信息（COM）")
        cover_sheet = wb_com.Sheets(1)  # 封面通常是第一个sheet

        # 搜索期数
        for row in range(1, 21):
            for col in range(5, 41):
                cell_val = cover_sheet.Cells(row, col).Value
                if cell_val and "第" in str(cell_val) and "期" in str(cell_val):
                    m = re.search(r'第(\d+)期', str(cell_val))
                    if m:
                        print(f"  [调试] 找到期数: {cover_sheet.Cells(row, col).Address} = {cell_val}")
                        cover_sheet.Cells(row, col).Value = f"第{issue_num}期"
                        print(f"  [成功] 修改期数: 第{issue_num}期")
                        break

        # 搜索天气
        weather_text = weather.split()[0] if weather else ""
        for row in range(1, 21):
            for col in range(5, 41):
                cell_val = cover_sheet.Cells(row, col).Value
                if cell_val:
                    for kw in WEATHER_KEYWORDS:
                        if kw in str(cell_val):
                            print(f"  [调试] 找到天气: {cover_sheet.Cells(row, col).Address} = {cell_val}")
                            if "天气：" in str(cell_val):
                                cover_sheet.Cells(row, col).Value = re.sub(r'天气：\S+', f'天气：{weather_text}', str(cell_val))
                            else:
                                cover_sheet.Cells(row, col).Value = weather_text
                            print(f"  [成功] 修改天气: {weather_text}")
                            break

        # ===== 步骤2: 处理日期列 =====
        print("\n[工具] 步骤2: 处理日期列（COM）")
        done = set()

        for sheet in TARGET_SHEETS:
            check_cancelled(cancel_check)
            if sheet in done:
                continue
            done.add(sheet)

            try:
                ws_com = wb_com.Sheets[sheet]
            except:
                print(f"  [跳过] 无法访问工作表: {sheet}")
                continue

            if ws_com.Visible == False:
                print(f"  [跳过] 工作表被隐藏: {sheet}")
                continue

            max_row = ws_com.UsedRange.Rows.Count
            max_col = ws_com.UsedRange.Columns.Count

            # 找日期列
            date_cols = []
            for col in range(1, max_col + 1):
                cell_val = ws_com.Cells(1, col).Value
                if cell_val and ('日期' in str(cell_val) or '20' in str(cell_val)):
                    date_cols.append(col)

            if len(date_cols) < 3:
                print(f"  [失败] {sheet}: 仅{len(date_cols)}个日期列，跳过")
                continue

            c1, c2, c3 = date_cols[:3]
            print(f"  [数据] {sheet} | 列: {[get_column_letter(c) for c in date_cols]}")

            # ===== 先更新第3列公式 =D1+days =====
            old_formula = ws_com.Cells(1, c3).Value
            ws_com.Cells(1, c3).Value = f"=D1+{days}"
            print(f"  [公式] 第{c3}列第1行: {old_formula} → =D1+{days}")

            # 等待Excel完成初始化计算
            time.sleep(2.0)
            ws_com.Calculate()
            time.sleep(1.0)

            # 数据迁移：第2列→第1列（触发重新计算）
            check_cancelled(cancel_check)
            c2_range = ws_com.Range(ws_com.Cells(1, c2), ws_com.Cells(max_row, c2))
            c1_range = ws_com.Range(ws_com.Cells(1, c1), ws_com.Cells(max_row, c1))
            c2_values = _range_to_column_values(c2_range.Value)
            c1_range.Value = _column_values_to_range(c2_values)
            print(f"  [列表] 第{c2}列→第{c1}列 ({max_row - 1}行，Range批量)")

            # 修改表头日期
            old_date = ws_com.Cells(1, c2).Value
            ws_com.Cells(1, c2).Value = target_date.strftime('%Y/%m/%d')
            print(f"  [日期] {old_date}→{target_date.strftime('%Y/%m/%d')}")

            # 等待计算完成
            ws_com.Calculate()
            time.sleep(1.0)

            # 第3列→第2列：只写数值，保持空值不覆盖的原逻辑
            check_cancelled(cancel_check)
            if max_row >= 2:
                c3_data_range = ws_com.Range(ws_com.Cells(2, c3), ws_com.Cells(max_row, c3))
                c2_data_range = ws_com.Range(ws_com.Cells(2, c2), ws_com.Cells(max_row, c2))
                c3_values = _range_to_column_values(c3_data_range.Value)
                old_c2_values = c2_values[1:]
                merged_values = []
                n2 = 0
                for idx, val in enumerate(c3_values):
                    if val is not None:
                        merged_values.append(val)
                        n2 += 1
                    else:
                        merged_values.append(old_c2_values[idx] if idx < len(old_c2_values) else None)
                c2_data_range.Value = _column_values_to_range(merged_values)
            else:
                n2 = 0
            print(f"  [列表] 第{c3}列→第{c2}列 ({n2}行，仅数值，Range批量）")

        check_cancelled(cancel_check)
        wb_com.Save()
        wb_com.Close()
        wb_com = None
        excel_com.Quit()
        excel_com = None
        _kill_excel_process()

        print(f"\n[成功] 已保存")
        return True
    except PipelineCancelled:
        print(f"[停止] 用户请求停止，正在关闭Excel: {filepath}")
        raise
    except Exception as e:
        print(f"[失败] 处理失败: {filepath}: {e}")
        return False
    finally:
        if wb_com is not None:
            try:
                wb_com.Close(SaveChanges=False)
            except Exception:
                pass
        if excel_com is not None:
            try:
                excel_com.Quit()
            except Exception:
                pass
        _kill_excel_process()

# ============ Excel转PDF ============
def excel_to_pdf(excel_path, pdf_path=None, max_retries=MAX_RETRIES, cancel_check=None):
    """使用Excel COM接口将Excel转换为PDF，保持所有格式。失败自动重试最多3次。"""
    if not os.path.exists(excel_path):
        print(f"[失败] Excel文件不存在: {excel_path}")
        return False

    if pdf_path is None:
        excel_dir = os.path.dirname(excel_path)
        excel_name = os.path.splitext(os.path.basename(excel_path))[0]
        pdf_path = os.path.join(excel_dir, f"{excel_name}.pdf")

    print(f"\n[文件] 正在转换: {os.path.basename(excel_path)}")
    print(f"   输出: {os.path.basename(pdf_path)}")

    last_error = None

    for attempt in range(max_retries):
        check_cancelled(cancel_check)
        excel = None
        wb = None
        try:
            if os.path.exists(pdf_path):
                print(f"[警告] PDF已存在，将被覆盖: {pdf_path}")

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = EXCEL_VISIBLE
            print(f"[DEBUG Step6] excel_type={EXCEL_TYPE}, EXCEL_VISIBLE={EXCEL_VISIBLE}, Visible set to {EXCEL_VISIBLE}")
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False

            check_cancelled(cancel_check)
            wb = excel.Workbooks.Open(excel_path)
            sheet_count = wb.Worksheets.Count
            print(f"   发现 {sheet_count} 个工作表")

            check_cancelled(cancel_check)
            wb.ExportAsFixedFormat(0, pdf_path)  # 0 = xlTypePDF

            if os.path.exists(pdf_path):
                size_kb = os.path.getsize(pdf_path) / 1024
                print(f"   [成功] PDF已生成: {size_kb:.1f} KB")
                return True
            else:
                print(f"   [失败] PDF生成失败（文件不存在）")

        except PipelineCancelled:
            print("   [停止] 用户请求停止PDF转换")
            raise
        except Exception as e:
            last_error = e
            if attempt < max_retries - 1:
                wait_seconds = RETRY_BACKOFF[attempt]
                print(f"   [重试] 第 {attempt + 1} 次失败: {e}，等待 {wait_seconds}s...")
                time.sleep(wait_seconds)
            else:
                print(f"   [最终] PDF转换失败（{max_retries}次重试）: {e}")

        finally:
            if wb is not None:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            _kill_excel_process()

    return False

def copy_directory(src, year, days, force=False):
    """复制并重命名目录，force=True时自动覆盖不询问"""
    if not os.path.exists(src):
        print(f"[失败] 源目录不存在: {src}")
        return None

    new_name, new_date = calculate_new_date(Path(src).name, year, days)
    if new_name is None:
        return None

    dst = Path(src).parent / new_name

    if dst.exists():
        if not force:
            response = input(f"[警告] 目录已存在 {dst}，覆盖？(y/n): ").lower()
            if response != 'y':
                print("操作取消")
                return None
        try:
            shutil.rmtree(dst)
        except Exception as e:
            print(f"[失败] 删除已有目录失败: {e}")
            return None

    skipped_temp_files = find_temp_excel_files(src)
    shutil.copytree(src, dst, ignore=_ignore_temp_excel_files)
    print(f"[成功] {src} -> {dst}")
    if skipped_temp_files:
        print(f"[清理] 复制时已跳过 {len(skipped_temp_files)} 个 Excel/WPS 临时锁文件")
    return str(dst), new_date

def delete_pdfs(directory):
    """删除目录下所有PDF"""
    count = 0
    for root, _, files in os.walk(directory):
        for f in files:
            if f.lower().endswith('.pdf'):
                try:
                    os.remove(os.path.join(root, f))
                    count += 1
                    print(f"[删除] 已删除: {f}")
                except Exception as e:
                    print(f"[失败] 删除失败: {f}: {e}")
    print(f"[数据] 共删除 {count} 个PDF")
    return count

# ============ 主流程 ============
def parse_user_input(raw):
    """解析用户输入，返回 (n_batches, days_list)"""
    parts = [x.strip() for x in raw.split(",")]
    if len(parts) < 2:
        raise ValueError("格式如：2,1,3（第1期间隔1天，第2期间隔3天）")
    n_batches = int(parts[0])
    days_list = [int(parts[i]) for i in range(1, len(parts))]
    return n_batches, days_list


def process_single_batch(source, latest_issue, batch, n_batches, days, weather_source=None, cancel_check=None):
    """
    处理单期：
    - 复制目录
    - 获取天气
    - 删除PDF
    - 重命名Excel
    - 处理汇总表
    - 处理外部引用
    - 转PDF
    返回 (target_dir, new_date, new_issue, weather, summary_dict)
    """
    next_issue_num = latest_issue + batch
    print("\n" + "=" * 56)
    print(f"     第 {batch}/{n_batches} 期  —  第{latest_issue + batch - 1}期 → 第{next_issue_num}期")
    print("=" * 56)

    print(f"\n[文件夹] 源目录: {source}")

    # 提取年份
    check_cancelled(cancel_check)
    excel_files = find_excel_files(source)
    if not excel_files:
        print("[失败] 未找到Excel文件，跳过本期")
        return None
    year = get_year_from_excel(excel_files[0])

    # 复制目录
    check_cancelled(cancel_check)
    result = copy_directory(source, year or datetime.now().year, days, force=True)
    if result is None:
        print("[失败] 复制目录失败，跳过本期")
        return None
    target_dir, new_date = result
    print(f"[成功] 新日期: {new_date.strftime('%Y/%m/%d')}")

    # 新期数
    issue, _, _ = parse_folder_info(Path(source).name)
    new_issue = issue + 1 if issue else None

    # 天气
    print("\n" + "=" * 28)
    print("        获取济南天气")
    print("=" * 28)
    check_cancelled(cancel_check)
    weather = get_weather(new_date)
    print(f"[成功] 获取到天气: {weather}")

    # 删除PDF
    check_cancelled(cancel_check)
    delete_pdfs(target_dir)

    # 重命名Excel文件（根据新期数）
    print("\n" + "=" * 28)
    print("        重命名Excel文件")
    print("=" * 28)
    check_cancelled(cancel_check)
    if issue and new_issue:
        rename_excel_files_by_issue(target_dir, issue, new_issue)
    else:
        print("[警告] 无法解析期数，跳过重命名")

    # 处理汇总表
    check_cancelled(cancel_check)
    files = find_excel_files(target_dir)
    if not files:
        print("[失败] 未找到汇总表，跳过本期")
        # 仍返回结果，让 caller 更新 current_source 到 target_dir
        return target_dir, new_date, new_issue, weather, None

    print(f"\n[文件夹] 找到 {len(files)} 个汇总表:")
    for i, f in enumerate(files, 1):
        print(f"  {i}. {os.path.basename(f)}")

    ok = 0
    for f in files:
        check_cancelled(cancel_check)
        if process_excel(f, new_date, new_issue, weather, days, cancel_check=cancel_check):
            ok += 1

    # 处理外部引用
    check_cancelled(cancel_check)
    handle_external_links_auto(target_dir, cancel_check=cancel_check)

    # Excel转PDF（只转第XX期.xlsx）
    print("\n" + "=" * 28)
    print("        Excel转PDF")
    print("=" * 28)

    pdf_ok = 0
    all_excels = find_all_excel_files(target_dir)
    for excel_path in all_excels:
        check_cancelled(cancel_check)
        basename = os.path.basename(excel_path)
        if re.search(r'第\d+期', basename):
            if excel_to_pdf(excel_path, cancel_check=cancel_check):
                pdf_ok += 1
        else:
            print(f"  [跳过] 跳过: {basename}")

    return target_dir, new_date, new_issue, weather, {
        "ok": ok,
        "total": len(files),
        "pdf_ok": pdf_ok,
    }


def find_latest_folder(parent_folder):
    """从父目录找出最新一期文件夹"""
    subfolders = [d for d in Path(parent_folder).iterdir() if d.is_dir()]
    issue_folders = [
        (parse_folder_info(d.name)[0], d)
        for d in subfolders
        if parse_folder_info(d.name)[0] is not None
    ]
    issue_folders.sort(key=lambda x: x[0])
    return issue_folders[-1] if issue_folders else (None, None)


def preflight_check(source_folder, excel_type=None):
    """运行前轻量自检，尽早发现目录和配置问题"""
    excel_type = (excel_type or EXCEL_TYPE).lower()
    issues = []
    warnings_list = []

    if excel_type not in ("office", "wps"):
        issues.append(f"Excel 类型无效: {excel_type}，应为 office 或 wps")

    if _desktop is None:
        warnings_list.append("pywinauto 桌面对象初始化失败，外部链接弹窗可能无法自动点击")

    try:
        import win32com.client as _win32com_client  # noqa: F401
    except Exception as e:
        issues.append(f"pywin32 / win32com 不可用: {e}")

    source_path = Path(source_folder)
    if not source_path.exists():
        issues.append(f"起始路径不存在: {source_path}")
        return False, None, issues, warnings_list

    if source_path.is_dir() and parse_folder_info(source_path.name)[0] is not None:
        parent = source_path.parent
    else:
        parent = source_path
    if not parent.exists():
        issues.append(f"父目录不存在: {parent}")
        return False, None, issues, warnings_list

    latest_issue, latest_folder = find_latest_folder(parent)
    if latest_folder is None:
        issues.append(f"未在父目录中找到形如 第43期 05.30 的期数文件夹: {parent}")
        return False, None, issues, warnings_list

    excel_files = find_excel_files(latest_folder)
    if not excel_files:
        issues.append(f"最新一期中未找到汇总表: {latest_folder}")

    temp_files = find_temp_excel_files(latest_folder)
    if temp_files:
        warnings_list.append(f"最新一期存在 {len(temp_files)} 个 ~$ 临时锁文件，程序会跳过且复制新目录时不再带入")

    return not issues, (latest_issue, latest_folder), issues, warnings_list


def print_preflight_result(ok, issues, warnings_list):
    """打印运行前自检结果"""
    print("\n[自检] 运行前检查")
    for item in warnings_list:
        print(f"  [提醒] {item}")
    if ok:
        print("  [通过] 基础环境和工程目录检查通过")
    else:
        for item in issues:
            print(f"  [失败] {item}")


def run_pipeline(source_folder, n_batches, days_list, progress_callback=None, excel_type=None, cancel_check=None):
    """
    给GUI用的入口函数，封装process_single_batch
    source_folder: 起始文件夹路径
    n_batches: 总期数
    days_list: 每期间隔天数列表
    excel_type: "office" 或 "wps"，动态设置 EXCEL_VISIBLE
    progress_callback(status_text, current, total) -> None
    cancel_check() -> bool，GUI请求停止时返回True
    返回: (success_count, fail_count, results_list)
    """
    # 动态设置 EXCEL_TYPE 和 EXCEL_VISIBLE（支持 GUI 传入）
    global EXCEL_TYPE, EXCEL_VISIBLE
    if excel_type is None:
        excel_type = os.environ.get("RPT_EXCEL_TYPE", "office")
    EXCEL_TYPE = excel_type
    EXCEL_VISIBLE = (EXCEL_TYPE.lower() == "wps")
    print(f"[DEBUG] run_pipeline excel_type={repr(excel_type)}, EXCEL_VISIBLE={EXCEL_VISIBLE}")

    ok, latest_info, issues, warnings_list = preflight_check(source_folder, excel_type=EXCEL_TYPE)
    print_preflight_result(ok, issues, warnings_list)
    if not ok:
        return 0, 1, [{"status": "fail", "msg": "运行前自检失败", "issues": issues}]

    latest_issue, latest_folder = latest_info

    results = []
    current_source = latest_folder

    try:
        for batch in range(1, n_batches + 1):
            check_cancelled(cancel_check)
            days = days_list[batch - 1] if batch - 1 < len(days_list) else 3

            if progress_callback:
                progress_callback(f"第{batch}/{n_batches}期 - 复制目录...", batch, n_batches)

            result = process_single_batch(
                str(current_source), latest_issue, batch, n_batches, days, cancel_check=cancel_check
            )

            if result is None:
                results.append({"status": "fail", "msg": "处理失败"})
                continue

            target_dir, new_date, new_issue, weather, summary = result
            results.append({
                "status": "ok",
                "target_dir": target_dir,
                "new_date": new_date.strftime('%Y/%m/%d'),
                "new_issue": new_issue,
                "weather": weather,
                "ok": summary["ok"] if summary else 0,
                "total": summary["total"] if summary else 0,
                "pdf_ok": summary["pdf_ok"] if summary else 0,
            })

            current_source = Path(target_dir)
    except PipelineCancelled as e:
        print(f"[停止] {e}")
        results.append({"status": "cancelled", "msg": str(e)})

    success = sum(1 for r in results if r["status"] == "ok")
    return success, len(results) - success, results


def main():
    base = r"C:\Users\Lenovo\Desktop"
    parent_folder_env = os.environ.get('REPORT_MAKER_PARENT')
    parent_folder = Path(parent_folder_env) if parent_folder_env else Path(base) / "新建文件夹"

    print("=" * 56)
    print("     Excel汇总表处理 + Excel转PDF 一条龙（批量）")
    print("=" * 56)

    # 找出当前最新的一期作为起始点
    ok, latest_info, issues, warnings_list = preflight_check(parent_folder, excel_type=EXCEL_TYPE)
    print_preflight_result(ok, issues, warnings_list)
    if not ok:
        return
    latest_issue, latest_folder = latest_info
    print(f"\n检测到最新一期: {latest_folder.name}")
    print(f"   将以此为起点制作新一期")

    # 输入
    raw = input("制作几期/各期间隔（如 2,1,3）: ").strip()
    try:
        n_batches, days_list = parse_user_input(raw)
    except ValueError as e:
        print(f"[失败] {e}")
        return

    print(f"\n将连续制作 {n_batches} 期（{latest_issue}期 → 第{latest_issue + n_batches}期）")

    current_source = latest_folder
    for batch in range(1, n_batches + 1):
        days = days_list[batch - 1] if batch - 1 < len(days_list) else 3
        result = process_single_batch(
            str(current_source), latest_issue, batch, n_batches, days
        )

        if result is None:
            # source 没有更新到 target_dir，保持原值
            continue

        target_dir, new_date, new_issue, weather, summary = result

        if summary:
            print("\n" + "=" * 56)
            print(f"         第 {batch}/{n_batches} 期完成摘要")
            print("=" * 56)
            print(f"[文件夹] 目标目录: {target_dir}")
            print(f"[日期] 新日期: {new_date.strftime('%Y/%m/%d')}")
            print(f"[数字] 新期数: 第{new_issue}期")
            print(f"[天气] 天气: {weather}")
            print(f"[数据] 汇总表处理: {summary['ok']}/{summary['total']}")
            print(f"[文件] PDF转换: {summary['pdf_ok']}")
            print("=" * 56)

        current_source = Path(target_dir)

    print("\n[完成] 全部完成！")

if __name__ == "__main__":
    main()
