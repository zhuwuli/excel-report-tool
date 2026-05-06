#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel汇总表自动处理工具 + Excel转PDF 一条龙
流程：
  1. 复制目录（期数+1，日期+自定义天数）
  2. 删除目录下所有PDF文件
  3. 修改Excel封面（期数、天气、温度）
  4. 处理Excel日期列（列迁移只写数值）
  5. 自动处理外部引用（注册表+自动点击更新）
  6. Excel转PDF（自动处理公式重算）
"""

import os
import re
import shutil
import sys
import time
import subprocess
import json
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import warnings
import requests
import random
import win32com.client
import winreg
import pywinauto
from pywinauto.timings import wait_until_passes

warnings.filterwarnings('ignore')

# ============ 可选依赖 ============
try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

# 检测 IDE 环境，IDE 控制台可能不识别 ANSI 颜色码
# 检测 PyCharm: PYCHARM_HOSTED=1 或 PYCHARM=True 或 sys.executable 包含 PyCharm
# 检测 VSCode: VSCODE_PID 或 VSCODE_INJECTION
PYCHARM_MARKER = 'PyCharm' in sys.executable or 'pycharm' in sys.executable.lower()
IDE_ENV = (
    os.environ.get('PYCHARM_HOSTED') is not None or
    os.environ.get('PYCHARM') is not None or
    os.environ.get('VSCODE_PID') is not None or
    os.environ.get('VSCODE_INJECTION') is not None or
    PYCHARM_MARKER
)

# IDE 环境强制禁用颜色（PyCharm/VSCode 等不识别 ANSI 码）
# 非 IDE 环境：只在 TTY 终端下启用颜色
if IDE_ENV or not sys.stdout.isatty():
    HAS_COLOR = False
else:
    try:
        import colorama
        colorama.init(autoreset=True)
        HAS_COLOR = True
    except ImportError:
        HAS_COLOR = False

def _c(text, code):
    """给文本加 ANSI 颜色码"""
    return f"\033[{code}m{text}\033[0m" if HAS_COLOR else text

def green(text):   return _c(text, "92")
def red(text):     return _c(text, "91")
def yellow(text):  return _c(text, "93")
def cyan(text):    return _c(text, "96")
def bold(text):    return _c(text, "1")

# ============ Windows Console 进度条（绕过 stdout 管道缓冲） ============
import ctypes

STD_OUTPUT_HANDLE = -11

class WindowsConsoleProgress:
    """用 Windows Console API 画进度条，绕过 Python stdout 管道缓冲"""

    def __init__(self, total, desc="处理中"):
        self.total = total
        self.current = 0
        self.desc = desc
        self._handle = ctypes.windll.kernel32.GetStdHandle(STD_OUTPUT_HANDLE)
        self._done = False

    def _write(self, text):
        """直接写 Windows Console，绕过 Python stdout 管道缓冲"""
        ctypes.windll.kernel32.WriteConsoleW(self._handle, text, len(text), None, None)

    def _clear_line(self):
        """清除当前行（ANSI 清除到行尾）"""
        self._write('\r' + ' ' * 70 + '\r')

    def update(self, current, label=""):
        self.current = current
        pct = int(current / self.total * 20) if self.total > 0 else 0
        bar = '█' * pct + '░' * (20 - pct)
        label_text = f" {label}" if label else ""
        msg = f"  {self.desc}: {bar} {current}/{self.total}{label_text}"
        self._clear_line()
        self._write(msg)

    def finish(self):
        if not self._done:
            self._write('\r' + ' ' * 70 + '\r')
            self._done = True

HAS_WIN_CONSOLE = sys.stdout.isatty()  # 仅在真终端启用，管道模式退化

# ============ 配置 ============

TARGET_SHEETS = {"垂直位移", "墙顶水平位移", "坑外水位", "测斜", "水平位移", "竖向位移", "轴力"}
WEATHER_API_KEY = "S-db8PJBJo-bZ0rb0" #心知天气的api key，过期记得更新

# 天气关键词
WEATHER_KEYWORDS = ["晴", "多云", "阴", "小雨", "中雨", "大雨", "雷阵雨", "小雪", "大雪", "雾", "霾"]

# 季节天气配置
SEASON_WEATHER = {
    (3, 4, 5): ["晴", "多云", "阴", "小雨"],
    (6, 7, 8): ["晴", "多云", "雷阵雨", "大雨"],
    (9, 10, 11): ["晴", "多云", "阴", "小雨"],
    (12, 1, 2): ["晴", "多云", "阴", "小雪"],
}

# 季节温度配置
SEASON_TEMP = {
    (12, 1, 2): (-5, 5),
    (3, 4, 5): (10, 20),
    (6, 7, 8): (25, 35),
    (9, 10, 11): (15, 25),
}

# 心知天气代码翻译
WEATHER_CODE_MAP = {
    "0": "晴", "1": "多云", "2": "阴", "3": "小雨", "4": "中雨",
    "5": "大雨", "6": "雷阵雨", "7": "小雪", "8": "中雪", "9": "大雪",
    "10": "雾", "13": "阵雨", "14": "雷阵雨", "19": "霾", "23": "雾",
}

# Open-Meteo WMO 天气码映射（主用/备用/历史共用）
WMO_CODE_MAP = {
    0: "晴", 1: "晴", 2: "多云", 3: "阴",
    45: "雾", 48: "雾", 51: "小雨", 53: "小雨", 55: "小雨",
    61: "雨", 63: "雨", 65: "雨",
    80: "阵雨", 81: "阵雨", 82: "阵雨",
    95: "雷阵雨", 96: "雷阵雨", 99: "雷阵雨"
}

# 济南坐标（Open-Meteo 用）
JINAN_LAT = 36.6512
JINAN_LON = 117.1205

# ============ 重试配置 ============
MAX_RETRIES = 3          # 各操作最大重试次数
RETRY_BACKOFF = [1, 2, 4]  # 指数退避间隔（秒）


# ============ 外部链接处理 ============
def add_trusted_location(folder_path):
    """将目录添加到Excel信任位置（注册表），避免黄色警告条"""
    try:
        key_path = r"Software\Microsoft\Office\16.0\Excel\Security\Trusted Locations"
        key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, key_path)
        sub_key_name = f"Location{abs(hash(folder_path)) % 100000}"
        sub_key = winreg.CreateKey(key, sub_key_name)
        winreg.SetValueEx(sub_key, "Path", 0, winreg.REG_SZ, folder_path)
        winreg.SetValueEx(sub_key, "AllowSubFolders", 0, winreg.REG_DWORD, 1)
        winreg.CloseKey(sub_key)
        winreg.CloseKey(key)
        print(f"  [成功] 已添加信任位置: {folder_path}")
        return True
    except Exception as e:
        print(f"  [失败] 添加信任位置失败: {e}")
        return False

def _kill_excel_process():
    """强制杀掉Excel进程，确保COM实例完全释放"""
    try:
        subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"],
                       capture_output=True, timeout=5)
        time.sleep(1.5)  # 等待进程完全退出
    except Exception:
        pass

_desktop = pywinauto.Desktop(backend="win32")

def _click_excel_update_button():
    """用pywinauto点击Excel'是否更新链接'对话框的'更新'按钮"""
    def try_click():
        try:
            for win in _desktop.windows():
                if win.window_text() == "Microsoft Excel":
                    for child in win.children():
                        txt = child.window_text()
                        # 精确匹配"更新(&U)"按钮
                        if txt == "更新(&U)":
                            child.click()
                            return True
        except Exception:
            pass
        return False

    try:
        wait_until_passes(12, 0.5, try_click)
        return True
    except Exception:
        return False

def open_and_update_links(excel_path, max_retries=MAX_RETRIES):
    """
    用COM打开Excel文件，自动处理'是否更新外部链接'对话框。
    流程：打开文件（UpdateLinks=2）→ 等对话框弹出 → pywinauto点'更新' → 保存关闭
    失败时最多重试 max_retries 次（指数退避）
    """
    last_error = None

    for attempt in range(max_retries):
        def _do():
            _kill_excel_process()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            time.sleep(2.0)

            wb = excel.Workbooks.Open(excel_path, UpdateLinks=2)
            time.sleep(2.0)

            clicked = _click_excel_update_button()
            if clicked:
                print(f"  [成功] 自动点击了'更新'按钮")
                time.sleep(1.0)  # 等 Excel 完成链接更新和计算
            else:
                print(f"  [提示] 未检测到链接更新对话框（可能无外部链接）")

            wb.Save()
            wb.Close()
            excel.Quit()
            _kill_excel_process()
            return True

        try:
            if attempt > 0:
                wait_seconds = RETRY_BACKOFF[attempt - 1]
                print(f"  [重试] 第 {attempt + 1} 次尝试，等待 {wait_seconds}s...")
                time.sleep(wait_seconds)
            result = _do()
            if result:
                return True
        except Exception as e:
            last_error = e
            print(f"  [失败] 第 {attempt + 1} 次: {e}")
            _kill_excel_process()

    print(f"  [最终] {os.path.basename(excel_path)} {max_retries}次重试均失败: {last_error}")
    return False

def handle_external_links_auto(target_dir):
    """
    自动处理目录下的外部引用：
    1. 将目录加入Excel信任位置（避免黄色警告条）
    2. 遍历所有Excel文件，每个文件单独开一个进程处理
    """
    print("\n[工具] 步骤5: 自动处理外部引用")

    # 1. 添加信任位置
    print("  [注册] 添加Excel信任位置...")
    add_trusted_location(target_dir)

# 2. 遍历所有Excel文件，自动处理外部链接
    all_excels = find_all_excel_files(target_dir)
    success_count = 0
    fail_count = 0

    # 用 Windows Console API 进度条，绕过 stdout 管道缓冲（实时更新）
    if HAS_WIN_CONSOLE:
        progress = WindowsConsoleProgress(len(all_excels), desc="Step5处理")
    else:
        progress = None

    for i, excel_path in enumerate(all_excels, 1):
        basename = os.path.basename(excel_path)
        if progress:
            progress.update(i, basename)
        else:
            print(f"  [处理] {basename}...", flush=True)

        if open_and_update_links(excel_path):
            success_count += 1
            print(f"  {green('[成功]')} {basename} 处理完成")
        else:
            fail_count += 1
            print(f"  {red('[失败]')} {basename} 处理失败（已重试{MAX_RETRIES}次）")
        time.sleep(1.0)

    if progress:
        progress.finish()
        print()  # 换行，避免后续输出覆盖同一行

    print(f"\n{cyan('[完成]')} 外部引用处理完成: {success_count}/{len(all_excels)} 个文件")
    if fail_count > 0:
        print(f"  {yellow('[警告]')} {fail_count} 个文件处理失败")
    return success_count



# ============ 工具函数 ============
def parse_folder_info(name):
    """解析文件夹名称，提取期数和日期"""
    m = re.search(r'第(\d+)期\s+(\d{2})\.(\d{2})', name)
    return (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else (None, None, None)

def calculate_new_date(folder_name, year, days):
    """计算新日期（期数+1，日期+days天）"""
    issue, month, day = parse_folder_info(folder_name)
    if issue is None:
        return None, None
    new_date = datetime(year, month, day) + timedelta(days=days)
    return f"第{issue + 1}期 {new_date.month:02d}.{new_date.day:02d}", new_date

def get_year_from_excel(path):
    """从Excel提取年份"""
    wb = None
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for col in range(1, 11):
            cell = ws.cell(row=1, column=col)
            if is_date_cell(cell):
                if isinstance(cell.value, datetime):
                    return cell.value.year
                else:
                    return datetime.strptime(str(cell.value), '%Y/%m/%d').year
        print("[警告] 未找到日期，使用当前年份")
        return None
    except Exception as e:
        print(f"[失败] 读取Excel失败: {e}")
        return None
    finally:
        if wb is not None:
            wb.close()

def get_weather(target_date):
    """获取天气（API优先，失败则模拟）"""
    days_diff = (target_date.date() - datetime.now().date()).days
    print(f"[搜索] 目标日期: {target_date.strftime('%Y-%m-%d')} (距今天 {days_diff} 天)")

    if days_diff < 0:
        # 过去日期，用 Historical API 查历史天气
        print(f"[调试] 过去日期，调用 Open-Meteo Historical...")
        w, t = fetch_open_meteo_historical(target_date)
        if w:
            print(f"[调试] Historical 返回: {w} {t}")
            return f"{w} {t}"
        print(f"[调试] Historical 失败，尝试 Open-Meteo Forecast...")
        w, t = fetch_open_meteo_weather()
        if w:
            return f"{w} {t}"
    elif days_diff <= 16:
        print(f"[调试] 尝试调用心知天气 API...")
        weather = fetch_api_weather(target_date, days_diff)
        if weather:
            print(f"[调试] 获取到天气: {weather}")
            return weather
    else:
        print(f"[调试] 日期超出 16 天范围，使用模拟天气")

    print("[切换] 使用模拟天气...")
    return make_simulated_weather(target_date)

def fetch_open_meteo_weather():
    """Open-Meteo Forecast API（备用，济南固定坐标），失败最多重试3次"""
    for attempt in range(MAX_RETRIES):
        try:
            url = (f"https://api.open-meteo.com/v1/forecast"
                   f"?latitude={JINAN_LAT}&longitude={JINAN_LON}"
                   f"&daily=weathercode,temperature_2m_max,temperature_2m_min"
                   f"&timezone=Asia/Shanghai&forecast_days=1")

            with urllib.request.urlopen(url, timeout=5) as r:
                data = json.loads(r.read().decode())

            code = data["daily"]["weathercode"][0]
            temp_max = data["daily"]["temperature_2m_max"][0]
            temp_min = data["daily"]["temperature_2m_min"][0]

            weather = WMO_CODE_MAP.get(code, "多云")
            temperature = f"{int(temp_min)}~{int(temp_max)}°C"
            return weather, temperature

        except Exception:
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_BACKOFF[attempt])
            else:
                return None, None
    return None, None

def fetch_open_meteo_historical(target_date):
    """Open-Meteo Historical API 获取历史天气（济南），失败最多重试3次"""
    for attempt in range(MAX_RETRIES):
        try:
            date_str = target_date.strftime('%Y-%m-%d')
            url = (f"https://archive-api.open-meteo.com/v1/archive"
                   f"?latitude={JINAN_LAT}&longitude={JINAN_LON}"
                   f"&start_date={date_str}&end_date={date_str}"
                   f"&daily=weathercode,temperature_2m_max,temperature_2m_min"
                   f"&timezone=Asia/Shanghai")

            with urllib.request.urlopen(url, timeout=10) as r:
                data = json.loads(r.read().decode())

            daily = data.get("daily", {})
            code = daily.get("weathercode", [None])[0]
            temp_max = daily.get("temperature_2m_max", [None])[0]
            temp_min = daily.get("temperature_2m_min", [None])[0]

            if code is None or temp_max is None:
                return None, None

            weather = WMO_CODE_MAP.get(code, "多云")
            temperature = f"{int(temp_min)}~{int(temp_max)}°C"

            print(f"  [历史] Open-Meteo Historical: {weather} {temperature}")
            return weather, temperature

        except Exception as e:
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_BACKOFF[attempt])
            else:
                print(f"  [错误] Open-Meteo Historical 失败: {e}")
                return None, None
    return None, None

def _fetch_seniverse_weather(target_date, days_diff):
    """从心知天气API获取天气（济南），失败最多重试3次"""
    for attempt in range(MAX_RETRIES):
        try:
            url = (f"https://api.seniverse.com/v3/weather/daily.json"
                   f"?key={WEATHER_API_KEY}&location=jinan&language=zh-Hans&unit=c"
                   f"&days={days_diff + 1}")
            resp = requests.get(url, timeout=10)
            if resp.status_code != 200:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_BACKOFF[attempt])
                    continue
                return None

            data = resp.json()
            daily = data.get("results", [{}])[0].get("daily", [])
            target_str = target_date.strftime('%Y-%m-%d')

            for day in daily:
                if day.get("date") == target_str:
                    text = day.get("text_day") or WEATHER_CODE_MAP.get(day.get("code_day", "1"), "多云")
                    low, high = day.get("low", "15"), day.get("high", "25")
                    result = f"{text} {low}~{high}°C"
                    print(f"[成功] API获取成功: {result}")
                    return result

            if daily:
                day = daily[0]
                text = WEATHER_CODE_MAP.get(day.get("code_day", "1"), "多云")
                low, high = day.get("low", "15"), day.get("high", "25")
                print(f"[警告] 日期匹配失败，使用最近可用数据")
                return f"{text} {low}~{high}°C"

        except Exception as e:
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_BACKOFF[attempt])
            else:
                print(f"[警告] API调用失败: {e}")
    return None

def fetch_api_weather(target_date, days_diff):
    """获取天气，先试心知天气，失败则用 Open-Meteo 备用"""
    # 心知天气只能覆盖约7天内，超过7天直接用 Open-Meteo
    if days_diff <= 2:
        weather = _fetch_seniverse_weather(target_date, days_diff)
        if weather is not None:
            return weather

    # 超出心知天气范围或心知天气失败，尝试 Open-Meteo
    print(f"  [备用] 尝试 Open-Meteo...")
    w, t = fetch_open_meteo_weather()
    if w:
        print(f"  [备用] Open-Meteo 天气: {w}, 温度: {t}")
        return f"{w} {t}"
    return None

def make_simulated_weather(target_date):
    """生成模拟天气"""
    month = target_date.month
    for months, weather_list in SEASON_WEATHER.items():
        if month in months:
            weather = random.choice(weather_list)
            low, high = SEASON_TEMP[months]
            return f"{weather} {low}~{high}°C"
    return "多云 15~25°C"

def find_excel_files(directory, patterns=None):
    """查找Excel文件"""
    patterns = patterns or ('汇总', 'summary')
    files = []
    for root, _, files_list in os.walk(directory):
        for f in files_list:
            if f.lower().endswith(('.xls', '.xlsx', '.xlsm')):
                if any(p.lower() in f.lower() for p in patterns):
                    files.append(os.path.join(root, f))
    return files

def find_all_excel_files(directory):
    """查找目录下所有Excel文件（不过滤文件名）"""
    return [
        os.path.join(directory, f)
        for f in os.listdir(directory)
        if f.lower().endswith(('.xls', '.xlsx', '.xlsm'))
    ]

def rename_excel_files_by_issue(directory, old_issue, new_issue):
    """根据期数重命名目录下的Excel文件（如 第87期.xlsx -> 第88期.xlsx）"""
    old_str = f"第{old_issue}期"
    new_str = f"第{new_issue}期"

    for root, _, files in os.walk(directory):
        for f in files:
            if old_str not in f:
                continue
            old_path = os.path.join(root, f)
            new_name = f.replace(old_str, new_str)
            new_path = os.path.join(root, new_name)
            try:
                os.rename(old_path, new_path)
                print(f"  [成功] 重命名: {f} -> {new_name}")
            except Exception as e:
                print(f"  [失败] 重命名失败: {f} -> {e}")

def is_date_cell(cell):
    """判断是否为日期单元格"""
    if not cell or cell.value is None:
        return False
    if isinstance(cell.value, datetime):
        return True
    fmt = str(cell.number_format).lower()
    return any(f in fmt for f in ['yy', 'yyyy', 'mm', 'dd', 'h', 'hh', 'mm:ss']) and \
        all(f not in fmt for f in ['$', '%', '0', '#', 'general'])

# ============ Excel处理 ============
def modify_cover(wb, issue_num, weather, date):
    """修改封面（期数、天气）"""
    try:
        ws = wb[wb.sheetnames[0]]
        weather_text = weather.split()[0]
        weather_ok = False

        def process_cell(cell):
            nonlocal weather_ok
            if cell.value is None:
                return
            val = str(cell.value)

            # 期数
            if "第" in val and "期" in val:
                m = re.search(r'第(\d+)期', val)
                if m:
                    print(f"  [调试] 找到期数单元格: {cell.coordinate} = {repr(cell.value)}")
                    new_val = re.sub(r'第\d+期', f'第{issue_num}期', val)
                    if new_val != val:
                        cell.value = new_val
                        print(f"  [成功] 修改期数: {cell.value}")

            # 天气关键词
            if not weather_ok:
                for kw in WEATHER_KEYWORDS:
                    if kw in val:
                        print(f"  [调试] 找到天气单元格: {cell.coordinate} = {repr(cell.value)}")
                        if "天气：" in val:
                            # 有前缀，保持前缀替换天气部分
                            cell.value = re.sub(r'天气：\S+', f'天气：{weather_text}', val)
                        else:
                            # 没有前缀，只替换天气关键词本身
                            cell.value = weather_text
                        print(f"  [成功] 修改天气: {cell.value}")
                        weather_ok = True
                        break

        # 封面区域搜索（1-20行，5-40列）
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=5, max_col=40):
            for cell in row:
                process_cell(cell)
                if weather_ok:
                    break

        # 全表搜索（期数和天气）
        if not weather_ok:
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=100):
                for cell in row:
                    process_cell(cell)

        print(f"[成功] 封面修改完成: 第{issue_num}期, 天气: {weather}")
        return True
    except Exception as e:
        print(f"[失败] 修改封面失败: {e}")
        return False

def process_excel(filepath, target_date, issue_num, weather):
    """处理汇总表"""
    if not os.path.exists(filepath):
        print(f"[失败] 文件不存在: {filepath}")
        return False

    try:
        wb = load_workbook(filepath)
        print(f"\n[文件] {os.path.basename(filepath)}")
        print(f"[日期] 目标日期: {target_date.strftime('%Y/%m/%d')}")

        # 封面
        print("\n[工具] 步骤1: 修改封面信息")
        modify_cover(wb, issue_num, weather, target_date)

        # 日期列处理
        print("\n[工具] 步骤2: 处理日期列")
        done = set()

        for sheet in wb.sheetnames:
            if sheet not in TARGET_SHEETS:
                print(f"[跳过] 跳过: {sheet}")
                continue
            if sheet in done:
                continue
            done.add(sheet)

            ws = wb[sheet]
            max_row, max_col = ws.max_row, ws.max_column

            # 找日期列
            date_cols = [col for col in range(1, max_col + 1) if is_date_cell(ws.cell(1, col))]

            if len(date_cols) < 3:
                print(f"  [失败] {sheet}: 仅{len(date_cols)}个日期列，跳过")
                continue

            c1, c2, c3 = date_cols[:3]
            print(f"  [数据] {sheet} | 列: {[get_column_letter(c) for c in date_cols]}")

            # 数据迁移：第2列→第1列
            n1 = sum(1 for r in range(1, max_row + 1) if ws.cell(r, c2).value)
            for r in range(1, max_row + 1):
                v = ws.cell(r, c2).value
                if v is not None:
                    ws.cell(r, c1).value = v
            print(f"  [列表] 第{c2}列→第{c1}列 ({n1}行)")

            old = ws.cell(1, c2).value
            ws.cell(1, c2).value = target_date.strftime('%Y/%m/%d')
            print(f"  [日期] {old}→{target_date.strftime('%Y/%m/%d')}")

            # 第3列→第2列：只写数值，不写公式
            # 用 data_only=True 重新打开，获取第3列的计算结果（不是公式）
            try:
                wb_data = load_workbook(filepath, data_only=True)
                ws_data = wb_data[sheet]
                n2 = 0
                for r in range(2, max_row + 1):
                    val = ws_data.cell(r, c3).value  # 这里是计算后的数值
                    if val is not None:
                        ws.cell(r, c2).value = val  # 直接写数值，不带公式
                        n2 += 1
                print(f"  [列表] 第{c3}列→第{c2}列 ({n2}行，仅数值）")
            finally:
                if wb_data is not None:
                    wb_data.close()

        wb.save(filepath)
        print(f"\n[成功] 已保存")
        return True
    except Exception as e:
        print(f"[失败] 处理失败: {filepath}: {e}")
        return False

# ============ Excel转PDF ============
def excel_to_pdf(excel_path, pdf_path=None, max_retries=MAX_RETRIES):
    """使用Excel COM接口将Excel转换为PDF，保持所有格式。失败自动重试最多3次。"""
    if not os.path.exists(excel_path):
        print(f"[失败] Excel文件不存在: {excel_path}")
        return False

    if pdf_path is None:
        excel_dir = os.path.dirname(excel_path)
        excel_name = os.path.splitext(os.path.basename(excel_path))[0]
        pdf_path = os.path.join(excel_dir, f"{excel_name}.pdf")

    print(f"\n[文件] 正在转换: {os.path.basename(excel_path)}")
    print(f"   输出: {os.path.basename(pdf_path)}")

    last_error = None

    for attempt in range(max_retries):
        excel = None
        wb = None
        try:
            if os.path.exists(pdf_path):
                print(f"[警告] PDF已存在，将被覆盖: {pdf_path}")

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False

            wb = excel.Workbooks.Open(excel_path)
            sheet_count = wb.Worksheets.Count
            print(f"   发现 {sheet_count} 个工作表")

            wb.ExportAsFixedFormat(0, pdf_path)  # 0 = xlTypePDF

            if os.path.exists(pdf_path):
                size_kb = os.path.getsize(pdf_path) / 1024
                print(f"   [成功] PDF已生成: {size_kb:.1f} KB")
                return True
            else:
                print(f"   [失败] PDF生成失败（文件不存在）")

        except Exception as e:
            last_error = e
            if attempt < max_retries - 1:
                wait_seconds = RETRY_BACKOFF[attempt]
                print(f"   [重试] 第 {attempt + 1} 次失败: {e}，等待 {wait_seconds}s...")
                time.sleep(wait_seconds)
            else:
                print(f"   [最终] PDF转换失败（{max_retries}次重试）: {e}")

        finally:
            if wb is not None:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            _kill_excel_process()

    return False

def copy_directory(src, year, days, force=False):
    """复制并重命名目录，force=True时自动覆盖不询问"""
    if not os.path.exists(src):
        print(f"[失败] 源目录不存在: {src}")
        return None

    new_name, new_date = calculate_new_date(Path(src).name, year, days)
    if new_name is None:
        return None

    dst = Path(src).parent / new_name

    if dst.exists():
        if not force:
            response = input(f"[警告] 目录已存在 {dst}，覆盖？(y/n): ").lower()
            if response != 'y':
                print("操作取消")
                return None
        try:
            shutil.rmtree(dst)
        except Exception as e:
            print(f"[失败] 删除已有目录失败: {e}")
            return None

    shutil.copytree(src, dst)
    print(f"[成功] {src} -> {dst}")
    return str(dst), new_date

def delete_pdfs(directory):
    """删除目录下所有PDF"""
    count = 0
    for root, _, files in os.walk(directory):
        for f in files:
            if f.lower().endswith('.pdf'):
                try:
                    os.remove(os.path.join(root, f))
                    count += 1
                    print(f"[删除] 已删除: {f}")
                except Exception as e:
                    print(f"[失败] 删除失败: {f}: {e}")
    print(f"[数据] 共删除 {count} 个PDF")
    return count

# ============ 主流程 ============
def main():
    base = r"C:\Users\Lenovo\Desktop\工作文件"
    parent_folder_env = os.environ.get('REPORT_MAKER_PARENT')
    if parent_folder_env:
        parent_folder = Path(parent_folder_env)
    else:
        parent_folder = Path(base) / "第87期 04.09"

    print("=" * 56)
    print("     Excel汇总表处理 + Excel转PDF 一条龙（批量）")
    print("=" * 56)

    # 找出当前最新的一期作为起始点
    subfolders = [d for d in parent_folder.iterdir() if d.is_dir()]
    issue_folders = []
    for d in subfolders:
        info = parse_folder_info(d.name)
        if info[0] is not None:
            issue_folders.append((info[0], d))
    issue_folders.sort(key=lambda x: x[0])
    latest_issue, latest_folder = issue_folders[-1]
    print(f"\n检测到最新一期: {latest_folder.name}")
    print(f"   将以此为起点制作新一期")

    # 输入：制作几期 + 各期间隔，如 "2,1,3" 表示2期，间隔1天和3天
    raw = input("制作几期/各期间隔（如 2,1,3）: ").strip()
    parts = [x.strip() for x in raw.split(",")]
    if len(parts) < 2:
        print("[失败] 格式如：2,1,3（第1期间隔1天，第2期间隔3天）")
        return
    try:
        n_batches = int(parts[0])
        days_list = [int(parts[i]) for i in range(1, len(parts))]
    except ValueError:
        print("[失败] 请输入数字，格式如：2,1,3")
        return

    print(f"\n将连续制作 {n_batches} 期（{latest_issue}期 → 第{latest_issue + n_batches}期）")

    current_source = latest_folder

    for batch in range(1, n_batches + 1):
        next_issue_num = latest_issue + batch
        days = days_list[batch - 1] if batch - 1 < len(days_list) else 3
        print("\n" + "=" * 56)
        print(f"     第 {batch}/{n_batches} 期  —  第{latest_issue + batch - 1}期 → 第{next_issue_num}期")
        print("=" * 56)

        source = str(current_source)
        print(f"\n[文件夹] 源目录: {source}")

        # 提取年份
        excel_files = find_excel_files(source)
        if not excel_files:
            print("[失败] 未找到Excel文件，跳过本期")
            continue
        year = get_year_from_excel(excel_files[0])

        # 复制目录
        result = copy_directory(source, year or datetime.now().year, days, force=True)
        if result is None:
            print("[失败] 复制目录失败，跳过本期")
            continue
        target_dir, new_date = result
        print(f"[成功] 新日期: {new_date.strftime('%Y/%m/%d')}")

        # 新期数
        issue, _, _ = parse_folder_info(Path(source).name)
        new_issue = issue + 1 if issue else None

        # 天气
        print("\n" + "=" * 28)
        print("        获取济南天气")
        print("=" * 28)
        weather = get_weather(new_date)
        print(f"[成功] 获取到天气: {weather}")

        # 删除PDF
        delete_pdfs(target_dir)

        # 重命名Excel文件（根据新期数）
        print("\n" + "=" * 28)
        print("        重命名Excel文件")
        print("=" * 28)
        if issue and new_issue:
            rename_excel_files_by_issue(target_dir, issue, new_issue)
        else:
            print("[警告] 无法解析期数，跳过重命名")

        # 处理汇总表
        files = find_excel_files(target_dir)
        if not files:
            print("[失败] 未找到汇总表，跳过本期")
            current_source = Path(target_dir)
            continue

        print(f"\n[文件夹] 找到 {len(files)} 个汇总表:")
        for i, f in enumerate(files, 1):
            print(f"  {i}. {os.path.basename(f)}")

        ok = sum(process_excel(f, new_date, new_issue, weather) for f in files)

        # 自动步骤：处理外部引用（注册表+pywinauto自动点击更新）
        handle_external_links_auto(target_dir)

        # Excel转PDF（只转第XX期.xlsx）
        print("\n" + "=" * 28)
        print("        Excel转PDF")
        print("=" * 28)

        pdf_ok = 0
        all_excels = find_all_excel_files(target_dir)
        for excel_path in all_excels:
            basename = os.path.basename(excel_path)
            if re.search(r'第\d+期', basename):
                if excel_to_pdf(excel_path):
                    pdf_ok += 1
            else:
                print(f"  [跳过] 跳过: {basename}")

        # 摘要
        print("\n" + "=" * 56)
        print(f"         第 {batch}/{n_batches} 期完成摘要")
        print("=" * 56)
        print(f"[文件夹] 目标目录: {target_dir}")
        print(f"[日期] 新日期: {new_date.strftime('%Y/%m/%d')}")
        print(f"[数字] 新期数: 第{new_issue}期")
        print(f"[天气] 天气: {weather}")
        print(f"[数据] 汇总表处理: {ok}/{len(files)}")
        print(f"[文件] PDF转换: {pdf_ok}")
        print("=" * 56)

        current_source = Path(target_dir)

    print("\n[完成] 全部完成！")

if __name__ == "__main__":
    main()