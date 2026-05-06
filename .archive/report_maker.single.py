#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel汇总表自动处理工具
功能：复制目录、重命名、删除PDF文件、处理汇总表Excel、自动修改封面
"""

import os
import re
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import warnings
import requests
import random

warnings.filterwarnings('ignore')


# ============ 配置 ============

TARGET_SHEETS = ["垂直位移", "墙顶水平位移", "坑外水位", "测斜", "水平位移", "竖向位移", "测斜", "轴力"]
WEATHER_API_KEY = "S-db8PJBJo-bZ0rb0"

# 天气关键词
WEATHER_KEYWORDS = ["晴", "多云", "阴", "小雨", "中雨", "大雨", "雷阵雨", "小雪", "大雪", "雾", "霾"]

# 季节天气配置
SEASON_WEATHER = {
    (3, 4, 5): ["晴", "多云", "阴", "小雨"],
    (6, 7, 8): ["晴", "多云", "雷阵雨", "大雨"],
    (9, 10, 11): ["晴", "多云", "阴", "小雨"],
    (12, 1, 2): ["晴", "多云", "阴", "小雪"],
}

# 季节温度配置
SEASON_TEMP = {
    (12, 1, 2): (-5, 5),
    (3, 4, 5): (10, 20),
    (6, 7, 8): (25, 35),
    (9, 10, 11): (15, 25),
}

# 心知天气代码翻译
WEATHER_CODE_MAP = {
    "0": "晴", "1": "多云", "2": "阴", "3": "小雨", "4": "中雨",
    "5": "大雨", "6": "雷阵雨", "7": "小雪", "8": "中雪", "9": "大雪",
    "10": "雾", "13": "阵雨", "14": "雷阵雨", "19": "霾", "23": "雾",
}


# ============ 工具函数 ============

def parse_folder_info(name):
    """解析文件夹名称，提取期数和日期"""
    m = re.search(r'第(\d+)期\s+(\d{2})\.(\d{2})', name)
    return (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else (None, None, None)


def calculate_new_date(folder_name, year, days):
    """计算新日期（期数+1，日期+days天）"""
    issue, month, day = parse_folder_info(folder_name)
    if issue is None:
        return None, None
    new_date = datetime(year, month, day) + timedelta(days=days)
    return f"第{issue + 1}期 {new_date.month:02d}.{new_date.day:02d}", new_date


def get_year_from_excel(path):
    """从Excel提取年份"""
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for col in range(1, 11):
            cell = ws.cell(row=1, column=col)
            if isinstance(cell.value, datetime):
                wb.close()
                print(f"📅 从Excel提取到年份: {cell.value.year}")
                return cell.value.year
        wb.close()
        print("⚠️ 未找到日期，使用当前年份")
        return None
    except Exception as e:
        print(f"❌ 读取Excel失败: {e}")
        return None


def get_weather(target_date):
    """获取天气（API优先，失败则模拟）"""
    days_diff = (target_date.date() - datetime.now().date()).days
    print(f"🔍 目标日期: {target_date.strftime('%Y-%m-%d')} (距今天 {days_diff} 天)")

    if 0 <= days_diff <= 15:
        weather = fetch_api_weather(target_date)
        if weather:
            return weather

    print("🔄 使用模拟天气...")
    return make_simulated_weather(target_date)


def fetch_api_weather(target_date):
    """从心知天气API获取天气"""
    try:
        url = f"https://api.seniverse.com/v3/weather/daily.json?key={WEATHER_API_KEY}&location=jinan&language=zh-Hans&unit=c&days={abs((target_date.date() - datetime.now().date()).days) + 1}"
        resp = requests.get(url, timeout=10)
        if resp.status_code != 200:
            return None

        data = resp.json()
        daily = data.get("results", [{}])[0].get("daily", [])
        target_str = target_date.strftime('%Y-%m-%d')

        for day in daily:
            if day.get("date") == target_str:
                text = day.get("text_day") or WEATHER_CODE_MAP.get(day.get("code_day", "1"), "多云")
                low, high = day.get("low", "15"), day.get("high", "25")
                result = f"{text} {low}~{high}°C"
                print(f"✅ API获取成功: {result}")
                return result

        if daily:
            day = daily[0]
            text = WEATHER_CODE_MAP.get(day.get("code_day", "1"), "多云")
            low, high = day.get("low", "15"), day.get("high", "25")
            print(f"⚠️ 日期匹配失败，使用最近可用数据")
            return f"{text} {low}~{high}°C"
    except Exception as e:
        print(f"⚠️ API调用失败: {e}")
    return None


def make_simulated_weather(target_date):
    """生成模拟天气"""
    month = target_date.month
    for months, weather_list in SEASON_WEATHER.items():
        if month in months:
            weather = random.choice(weather_list)
            low, high = SEASON_TEMP[months]
            return f"{weather} {low}~{high}°C"
    return "多云 15~25°C"


def find_excel_files(directory, patterns=None):
    """查找Excel文件"""
    patterns = patterns or ('汇总', 'summary')
    files = []
    for root, _, files_list in os.walk(directory):
        for f in files_list:
            if f.lower().endswith(('.xls', '.xlsx', '.xlsm')):
                if any(p.lower() in f.lower() for p in patterns):
                    files.append(os.path.join(root, f))
    return files


def rename_excel_files_by_issue(directory, old_issue, new_issue):
    """根据期数重命名目录下的Excel文件（如 第87期.xlsx -> 第88期.xlsx）"""
    renamed = []
    old_str = f"第{old_issue}期"
    new_str = f"第{new_issue}期"

    for root, _, files in os.walk(directory):
        for f in files:
            if old_str in f:
                old_path = os.path.join(root, f)
                new_name = f.replace(old_str, new_str)
                new_path = os.path.join(root, new_name)
                try:
                    os.rename(old_path, new_path)
                    renamed.append((f, new_name))
                    print(f"  ✅ 重命名: {f} -> {new_name}")
                except Exception as e:
                    print(f"  ❌ 重命名失败: {f} -> {e}")
    return renamed


def is_date_cell(cell):
    """判断是否为日期单元格"""
    if not cell or cell.value is None:
        return False
    if isinstance(cell.value, datetime):
        return True
    fmt = str(cell.number_format).lower()
    return any(f in fmt for f in ['yy', 'yyyy', 'mm', 'dd', 'h', 'hh', 'mm:ss']) and \
           all(f not in fmt for f in ['$', '%', '0', '#', 'general'])


# ============ Excel处理 ============

def modify_cover(wb, issue_num, weather, date):
    """修改封面（期数、天气，温度）"""
    try:
        ws = wb[wb.sheetnames[0]]
        weather_text = weather.split()[0]
        temp_text = weather.split()[-1]
        weather_ok = temp_ok = False

        def process_cell(cell):
            nonlocal weather_ok, temp_ok
            if not isinstance(cell.value, str):
                return
            val = cell.value

            # 期数
            if "第" in val and "期" in val:
                print(f"  [调试] 找到期数单元格: {cell.coordinate} = {repr(cell.value)}")
                new_val = re.sub(r'第\d+期', f'第{issue_num}期', val)
                if new_val != val:
                    cell.value = new_val
                    print(f"  ✅ 修改期数: {cell.value}")

            # 天气关键词
            if not weather_ok:
                for kw in WEATHER_KEYWORDS:
                    if kw in val:
                        print(f"  [调试] 找到天气单元格: {cell.coordinate} = {repr(cell.value)}")
                        cell.value = weather_text
                        print(f"  ✅ 修改天气: {cell.value}")
                        weather_ok = True
                        break

            # 温度（兼容范围和单值格式）
            if not temp_ok and re.search(r'\d+°C', val):
                cell.value = re.sub(r'\d+°C', temp_text, val)
                print(f"  ✅ 修改温度: {cell.value}")
                temp_ok = True

        # 封面区域搜索（1-20行，1-25列）
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=5, max_col=40):
            for cell in row:
                process_cell(cell)
                if weather_ok and temp_ok:
                    break

        # 备用：全表搜索（温度）
        if not temp_ok:
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=100):
                for cell in row:
                    if isinstance(cell.value, str) and re.search(r'\d+°C', cell.value):
                        cell.value = re.sub(r'\d+°C', temp_text, cell.value)
                        print(f"  ✅ (备用)修改温度: {cell.value}")
                        temp_ok = True
                        break
                if temp_ok:
                    break

        # 全表搜索（期数和天气）
        if not (weather_ok and temp_ok):
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=100):
                for cell in row:
                    process_cell(cell)
        print(f"✅ 封面修改完成: 第{issue_num}期, 天气: {weather}")
        return True
    except Exception as e:
        print(f"❌ 修改封面失败: {e}")
        return False


def process_excel(filepath, target_date, issue_num, weather):
    """处理汇总表"""
    if not os.path.exists(filepath):
        print(f"❌ 文件不存在: {filepath}")
        return False

    try:
        wb = load_workbook(filepath)
        print(f"\n📄 {os.path.basename(filepath)}")
        print(f"📅 目标日期: {target_date.strftime('%Y/%m/%d')}")

        # 封面
        print("\n🔧 步骤1: 修改封面信息")
        modify_cover(wb, issue_num, weather, target_date)

        # 日期列处理
        print("\n🔧 步骤2: 处理日期列")
        done = set()

        for sheet in wb.sheetnames:
            if sheet not in TARGET_SHEETS:
                print(f"⏭️ 跳过: {sheet}")
                continue
            if sheet in done:
                continue
            done.add(sheet)

            ws = wb[sheet]
            max_row, max_col = ws.max_row, ws.max_column

            # 找日期列
            date_cols = [col for col in range(1, max_col + 1) if is_date_cell(ws.cell(1, col))]

            if len(date_cols) < 3:
                print(f"  ❌ {sheet}: 仅{len(date_cols)}个日期列，跳过")
                continue

            c1, c2, c3 = date_cols[:3]
            print(f"  📊 {sheet} | 列: {[get_column_letter(c) for c in date_cols]}")

            # 数据迁移：第2列→第1列
            n1 = sum(1 for r in range(1, max_row + 1) if ws.cell(r, c2).value)
            for r in range(1, max_row + 1):
                v = ws.cell(r, c2).value
                if v is not None:
                    ws.cell(r, c1).value = v
            print(f"  📋 第{c2}列→第{c1}列 ({n1}行)")

            old = ws.cell(1, c2).value
            ws.cell(1, c2).value = target_date.strftime('%Y/%m/%d')
            print(f"  📅 {old}→{target_date.strftime('%Y/%m/%d')}")

            # 第3列→第2列：关键改动——只写数值，不写公式
            # 用 data_only=True 重新打开，获取第3列的计算结果（不是公式）
            wb_data = load_workbook(filepath, data_only=True)
            ws_data = wb_data[sheet]
            n2 = 0
            for r in range(2, max_row + 1):
                val = ws_data.cell(r, c3).value  # 这里是计算后的数值
                if val is not None:
                    ws.cell(r, c2).value = val     # 直接写数值，不带公式
                    n2 += 1
            wb_data.close()
            print(f"  📋 第{c3}列→第{c2}列 ({n2}行，仅数值）")

        wb.save(filepath)
        print(f"\n✅ 已保存")
        return True
    except Exception as e:
        print(f"❌ 处理失败: {filepath}: {e}")
        return False


# ============ 目录操作 ============

def copy_directory(src, year, days):
    """复制并重命名目录"""
    if not os.path.exists(src):
        print(f"❌ 源目录不存在: {src}")
        return None

    new_name, new_date = calculate_new_date(Path(src).name, year, days)
    if new_name is None:
        return None

    dst = Path(src).parent / new_name

    if dst.exists():
        if input(f"⚠️ 目录已存在 {dst}，覆盖？(y/n): ").lower() != 'y':
            print("操作取消")
            return None
        shutil.rmtree(dst)

    shutil.copytree(src, dst)
    print(f"✅ {src} -> {dst}")
    return str(dst), new_date


def delete_pdfs(directory):
    """删除目录下所有PDF"""
    count = 0
    for root, _, files in os.walk(directory):
        for f in files:
            if f.lower().endswith('.pdf'):
                try:
                    os.remove(os.path.join(root, f))
                    count += 1
                    print(f"🗑️ 已删除: {f}")
                except Exception as e:
                    print(f"❌ 删除失败: {f}: {e}")
    print(f"📊 共删除 {count} 个PDF")
    return count


# ============ 主流程 ============

def main():
    source = r"C:\Users\Lenovo\Desktop\工作文件\第2期 04.09"

    print("=" * 56)
    print("          Excel汇总表自动处理工具")
    print("=" * 56)
    print(f"\n📁 源目录: {source}")

    # 提取年份
    excel_files = find_excel_files(source)
    if not excel_files:
        print("❌ 未找到Excel文件")
        return
    year = get_year_from_excel(excel_files[0])

    # 输入天数
    days = int(input("请输入日期偏移天数（默认3天）: ").strip() or 3)

    # 复制目录
    result = copy_directory(source, year or datetime.now().year, days)
    if result is None:
        return
    target_dir, new_date = result
    print(f"✅ 新日期: {new_date.strftime('%Y/%m/%d')}")

    # 新期数
    issue, _, _ = parse_folder_info(Path(source).name)
    new_issue = issue + 1 if issue else None

    # 天气
    print("\n" + "=" * 28)
    print("        获取济南天气")
    print("=" * 28)
    weather = get_weather(new_date)
    print(f"✅ 获取到天气: {weather}")

    # 删除PDF
    delete_pdfs(target_dir)

    # 重命名Excel文件（根据新期数）
    print("\n" + "=" * 28)
    print("        重命名Excel文件")
    print("=" * 28)
    if issue and new_issue:
        rename_excel_files_by_issue(target_dir, issue, new_issue)
    else:
        print("⚠️ 无法解析期数，跳过重命名")

    # 处理汇总表
    files = find_excel_files(target_dir)
    if not files:
        print("❌ 未找到汇总表")
        return

    print(f"\n📁 找到 {len(files)} 个汇总表:")
    for i, f in enumerate(files, 1):
        print(f"  {i}. {os.path.basename(f)}")

    ok = sum(process_excel(f, new_date, new_issue, weather) for f in files)

    # 摘要
    print("\n" + "=" * 56)
    print("              处理完成摘要")
    print("=" * 56)
    print(f"📁 目标目录: {target_dir}")
    print(f"📅 新日期: {new_date.strftime('%Y/%m/%d')}")
    print(f"🔢 新期数: 第{new_issue}期")
    print(f"🌤️ 天气: {weather}")
    print(f"📊 已处理: {ok}/{len(files)}")
    print(f"⏰ 处理时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 56)


if __name__ == "__main__":
    try:
        from openpyxl import load_workbook
        import requests
    except ImportError as e:
        print(f"❌ 缺少依赖: {e}\n请运行: pip install openpyxl requests")
        exit(1)

    main()